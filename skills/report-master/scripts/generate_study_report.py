#!/usr/bin/env python3
"""
报告老油条 - 学习报告生成器
严格还原模板格式，填充老油条内容
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import sys
import os

def create_study_report(input_path, output_path, name="吕进", dept="股票学习机业务线", date="2026年3月24日"):
    """根据模板生成学习报告"""
    
    # 读取模板获取格式信息
    src_wb = openpyxl.load_workbook(input_path)
    src_ws = src_wb.active
    
    # 创建新工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学习报告"
    
    # 复制列宽
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
        src_dim = src_ws.column_dimensions[col_letter]
        ws.column_dimensions[col_letter].width = src_dim.width or 15.0
    
    # 定义边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 定义字体
    title_font = Font(name='微软雅黑', size=16, bold=True)
    normal_font = Font(name='Calibri', size=11)
    
    # 定义对齐
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # ===== 第1行：标题 =====
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = src_ws.row_dimensions[1].height or 30
    cell = ws['A1']
    cell.value = '公开课学习说明'
    cell.font = title_font
    cell.alignment = center_align
    cell.border = thin_border
    for c in ['B1','C1','D1','E1','F1']:
        ws[c].border = thin_border
    
    # ===== 第2行：欢迎语 =====
    ws.merge_cells('A2:F2')
    ws.row_dimensions[2].height = src_ws.row_dimensions[2].height or 45
    cell = ws['A2']
    cell.value = '亲爱的同学，你好！\n恭喜获得公开课学习机会，期待你的学习成长，并带回真知灼见，转化为内部价值'
    cell.font = normal_font
    cell.alignment = left_align
    cell.border = thin_border
    for c in ['B2','C2','D2','E2','F2']:
        ws[c].border = thin_border
    
    # ===== 第3-4行：课前思考 =====
    ws.merge_cells('A3:A4')
    cell = ws['A3']
    cell.value = '课前思考'
    cell.font = normal_font
    cell.alignment = center_align
    cell.border = thin_border
    
    ws.merge_cells('B3:F3')
    cell = ws['B3']
    cell.value = '请带着与该主题相关的问题参训，请在这里写下思考。'
    cell.font = normal_font
    cell.alignment = left_align
    cell.border = thin_border
    for c in ['C3','D3','E3','F3']:
        ws[c].border = thin_border
    
    ws.merge_cells('B4:F4')
    ws.row_dimensions[4].height = src_ws.row_dimensions[4].height or 100
    cell = ws['B4']
    cell.value = ('1. AI 智能体如何赋能投教产品，提升用户学习效果和转化率？\n'
                   '2. 股票学习机的软硬件结合场景，如何通过 AI 实现个性化学习路径？\n'
                   '3. ERP 系统如何支持 AI 驱动的用户行为分析和精准营销？\n'
                   '4. 如何将 AI 长期记忆能力应用到股民学习成长追踪中？\n'
                   '5. 智能体在金融合规前提下的应用边界在哪里？')
    cell.font = normal_font
    cell.alignment = left_align
    cell.border = thin_border
    for c in ['C4','D4','E4','F4']:
        ws[c].border = thin_border
    for c in ['A4']:
        ws[c].border = thin_border
    
    # ===== 第5-6行：课中记录 =====
    ws.merge_cells('A5:A6')
    cell = ws['A5']
    cell.value = '课中记录'
    cell.font = normal_font
    cell.alignment = center_align
    cell.border = thin_border
    
    ws.merge_cells('B5:F5')
    cell = ws['B5']
    cell.value = '请记录对你有用的课程内容，可以在该表格中填写，也可以是手写/思维导图/PPT 等格式的资料。'
    cell.font = normal_font
    cell.alignment = left_align
    cell.border = thin_border
    for c in ['C5','D5','E5','F5']:
        ws[c].border = thin_border
    
    ws.merge_cells('B6:F6')
    ws.row_dimensions[6].height = src_ws.row_dimensions[6].height or 180
    cell = ws['B6']
    cell.value = ('【核心收获】\n\n'
                   '一、AI 智能体趋势认知\n'
                   '• 2026 年 AI 智能体从"工具"向"伙伴"演进，具备长期记忆和上下文理解能力\n'
                   '• 智能体需要"调教→共生"的进化路径，与用户形成协作关系\n'
                   '• 群体智能和个体智能是两大发展方向\n\n'
                   '二、关键技术应用\n'
                   '• 记忆革命：AI 代理的长期记忆能力可应用于用户学习轨迹追踪\n'
                   '• 个性化表达：AI 时代的品牌重塑，投教内容需要个性化呈现\n'
                   '• 从工具到伙伴：股票学习机应从"学习工具"升级为"投资伙伴"\n\n'
                   '三、产业融合洞察\n'
                   '• AGI 与跨界融合：金融投教可与 AI 深度结合\n'
                   '• 可见性营销：从流量获取到商业转化的全链路\n'
                   '• 内容 OPC：AI 创作到变现的运营增长方法论\n\n'
                   '四、产品出海路径\n'
                   '• AI 产品的本地验证到全球扩张路径\n'
                   '• 软硬件结合的 ERP 系统可借鉴出海经验')
    cell.font = normal_font
    cell.alignment = left_align
    cell.border = thin_border
    for c in ['C6','D6','E6','F6']:
        ws[c].border = thin_border
    for c in ['A6']:
        ws[c].border = thin_border
    
    # ===== 第7行：课后转化（表头）=====
    ws.merge_cells('A7:A11')
    cell = ws['A7']
    cell.value = '课后转化'
    cell.font = normal_font
    cell.alignment = center_align
    cell.border = thin_border
    
    headers = [
        ('B7', '能在实际工作中应用的知识点'),
        ('C7', '计划\n（准备如何应用于工作？什么时候完成？）'),
        ('D7', '实施\n（具体操作步骤/怎么做？）'),
        ('E7', '检查\n（直属领导来监督）'),
        ('F7', '评价\n（应用过程中的自我心得体会）'),
    ]
    for coord, val in headers:
        cell = ws[coord]
        cell.value = val
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # ===== 第8-11行：课后转化内容 =====
    rows_data = [
        {
            'knowledge': 'AI 长期记忆能力应用到股票学习机用户成长追踪',
            'plan': '2026 年 Q2 完成方案设计，Q3 实现 MVP 版本',
            'do': '1. 在 ERP 系统中增加用户学习记忆模块\n2. 记录用户学习历史、错题本、投资决策轨迹\n3. 基于记忆数据生成个性化学习建议\n4. 实现"越用越懂你"的智能投教体验',
            'check': '由部门总监月度 Review 进度',
            'evaluate': '（待实施后填写）',
        },
        {
            'knowledge': '智能体"从调教到共生"理念应用到产品交互设计',
            'plan': '2026 年 Q2 完成产品原型设计',
            'do': '1. 重构股票学习机的用户引导流程\n2. 设计 AI 助手与用户的协作式学习场景\n3. 减少"调教"感，增强"伙伴"感\n4. 在硬件端增加语音/自然交互能力',
            'check': '由产品负责人双周检查',
            'evaluate': '（待实施后填写）',
        },
        {
            'knowledge': 'ERP 系统支持 AI 驱动的精准营销和转化',
            'plan': '2026 年 Q3 完成数据中台改造',
            'do': '1. 整合用户学习数据、硬件使用数据、购买行为数据\n2. 建立用户画像和转化预测模型\n3. 实现个性化课程推荐和硬件升级建议\n4. 通过 AI 优化营销触达时机和内容',
            'check': '由技术总监月度 Review',
            'evaluate': '（待实施后填写）',
        },
        {
            'knowledge': '内容 OPC 方法论应用到投教课程生产',
            'plan': '2026 年 Q2 启动试点',
            'do': '1. 用 AI 辅助投教课程内容的批量生产\n2. 建立从内容创作到用户变现的全链路追踪\n3. 测试 AI 生成内容 vs 人工内容的转化效果\n4. 优化内容 ROI 评估体系',
            'check': '由运营负责人双周检查',
            'evaluate': '（待实施后填写）',
        },
    ]
    
    for i, row_data in enumerate(rows_data):
        row_num = 8 + i
        ws.row_dimensions[row_num].height = src_ws.row_dimensions[row_num].height or 100
        
        ws[f'B{row_num}'].value = row_data['knowledge']
        ws[f'B{row_num}'].font = normal_font
        ws[f'B{row_num}'].alignment = left_align
        ws[f'B{row_num}'].border = thin_border
        
        ws[f'C{row_num}'].value = row_data['plan']
        ws[f'C{row_num}'].font = normal_font
        ws[f'C{row_num}'].alignment = left_align
        ws[f'C{row_num}'].border = thin_border
        
        ws[f'D{row_num}'].value = row_data['do']
        ws[f'D{row_num}'].font = normal_font
        ws[f'D{row_num}'].alignment = left_align
        ws[f'D{row_num}'].border = thin_border
        
        ws[f'E{row_num}'].value = row_data['check']
        ws[f'E{row_num}'].font = normal_font
        ws[f'E{row_num}'].alignment = left_align
        ws[f'E{row_num}'].border = thin_border
        
        ws[f'F{row_num}'].value = row_data['evaluate']
        ws[f'F{row_num}'].font = normal_font
        ws[f'F{row_num}'].alignment = left_align
        ws[f'F{row_num}'].border = thin_border
        
        ws[f'A{row_num}'].border = thin_border
    
    # ===== 第12行：学员信息 =====
    ws.row_dimensions[12].height = src_ws.row_dimensions[12].height or 25
    
    ws['A12'].value = '学员姓名'
    ws['A12'].font = normal_font
    ws['A12'].alignment = center_align
    ws['A12'].border = thin_border
    
    ws['B12'].value = name
    ws['B12'].font = normal_font
    ws['B12'].alignment = center_align
    ws['B12'].border = thin_border
    
    ws['C12'].value = '部门'
    ws['C12'].font = normal_font
    ws['C12'].alignment = center_align
    ws['C12'].border = thin_border
    
    ws['D12'].value = dept
    ws['D12'].font = normal_font
    ws['D12'].alignment = center_align
    ws['D12'].border = thin_border
    
    ws['E12'].value = '培训日期'
    ws['E12'].font = normal_font
    ws['E12'].alignment = center_align
    ws['E12'].border = thin_border
    
    ws['F12'].value = date
    ws['F12'].font = normal_font
    ws['F12'].alignment = center_align
    ws['F12'].border = thin_border
    
    # 保存
    wb.save(output_path)
    print(f"✅ 报告已生成：{output_path}")
    return output_path


if __name__ == "__main__":
    input_file = sys.argv[1] if len(sys.argv) > 1 else None
    output_file = sys.argv[2] if len(sys.argv) > 2 else "学习报告_输出.xlsx"
    
    if input_file and os.path.exists(input_file):
        create_study_report(input_file, output_file)
    else:
        # 直接使用默认数据生成
        create_study_report(
            "/home/oc/.openclaw/media/inbound/å_ä¹_æ_å_å_è_AI_æ_ç_ç_å³_ä¼---fa826719-f8fc-43c6-9142-93c8cb6c16bf.xlsx",
            output_file
        )
